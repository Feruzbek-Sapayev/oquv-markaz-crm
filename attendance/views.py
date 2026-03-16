import json
import calendar
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST

from .models import AttendanceSession, Attendance
from .forms import AttendanceSessionForm, AttendanceFormSet
from courses.models import Group, Enrollment
from accounts.permissions import admin_required, teacher_required
from config.telegram_utils import send_telegram_message
from notifications.models import Notification


UZ_MONTHS = [
    (1, 'yan'), (2, 'fev'), (3, 'mar'), (4, 'apr'), (5, 'may'), (6, 'iun'),
    (7, 'iul'), (8, 'avg'), (9, 'sen'), (10, 'okt'), (11, 'noy'), (12, 'dek')
]


@login_required
def session_list(request):
    group_id = request.GET.get('group', '')
    month_year = request.GET.get('month_year')
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    today = timezone.now().date()
    
    if month_year:
        try:
            month, year = map(int, month_year.split('-'))
        except (ValueError, TypeError):
            month, year = today.month, today.year
    else:
        try:
            month = int(month) if month else today.month
            year = int(year) if year else today.year
        except (ValueError, TypeError):
            month, year = today.month, today.year
        
    start_date = datetime.date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = datetime.date(year, month, last_day)
    
    # Calculate prev/next month
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year
        
    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    groups = Group.objects.filter(is_active=True).select_related('course')
    
    if request.user.is_teacher:
        groups = groups.filter(teacher__user=request.user).distinct()
    elif request.user.is_student:
        groups = groups.filter(enrollments__student__user=request.user, enrollments__is_active=True)
    elif not request.user.is_admin_role:
        groups = groups.none()
        
    selected_group = None
    matrix_data = []
    all_dates = []
    
    if group_id:
        selected_group = groups.filter(pk=group_id).first()
        
    if selected_group:
        potential_dates = [start_date + datetime.timedelta(days=i) for i in range(last_day)]
        schedule = selected_group.days
        if schedule == Group.DayChoices.ODD:
            all_dates = [d for d in potential_dates if d.weekday() in [0, 2, 4]]
        elif schedule == Group.DayChoices.EVEN:
            all_dates = [d for d in potential_dates if d.weekday() in [1, 3, 5]]
        elif schedule == Group.DayChoices.WEEKEND:
            all_dates = [d for d in potential_dates if d.weekday() in [5, 6]]
        else:
            all_dates = potential_dates
            
        enrollments = Enrollment.objects.filter(group=selected_group, is_active=True).select_related('student').order_by('student__last_name')
        
        attendance_records = Attendance.objects.filter(
            session__group=selected_group, session__date__range=[start_date, end_date]
        ).values('student_id', 'session__date', 'status')
        
        attendance_map = {}
        for r in attendance_records:
            sid, d, s = r['student_id'], r['session__date'], r['status']
            if sid not in attendance_map: attendance_map[sid] = {}
            attendance_map[sid][d] = s
            
        for e in enrollments:
            student_atts = []
            present_count = 0
            total_with_status = 0
            for d in all_dates:
                status = attendance_map.get(e.student.pk, {}).get(d)
                student_atts.append({'date': d, 'status': status})
                if status:
                    total_with_status += 1
                    if status == Attendance.Status.PRESENT:
                        present_count += 1
            
            percent = (present_count / total_with_status * 100) if total_with_status else 0
            
            if request.user.is_student:
                if e.student.user == request.user:
                    matrix_data.append({
                        'student': e.student,
                        'attendances': student_atts,
                        'percentage': round(percent),
                    })
            else:
                matrix_data.append({
                    'student': e.student,
                    'attendances': student_atts,
                    'percentage': round(percent),
                })

    total_present = 0
    total_absent = 0
    total_records = 0
    
    for row in matrix_data:
        for att in row['attendances']:
            if att['status'] == Attendance.Status.PRESENT:
                total_present += 1
                total_records += 1
            elif att['status'] == Attendance.Status.ABSENT:
                total_absent += 1
                total_records += 1
                
    overall_percentage = round((total_present / total_records * 100)) if total_records > 0 else 0

    return render(request, 'attendance/session_list.html', {
        'groups': groups,
        'selected_group': group_id,
        'selected_group_obj': selected_group,
        'matrix_data': matrix_data,
        'all_dates': all_dates,
        'current_month': month,
        'current_year': year,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'uz_months': UZ_MONTHS,
        'full_uz_months': [
            (1, 'Yanvar'), (2, 'Fevral'), (3, 'Mart'), (4, 'Aprel'), (5, 'May'), (6, 'Iyun'),
            (7, 'Iyul'), (8, 'Avgust'), (9, 'Sentabr'), (10, 'Oktabr'), (11, 'Noyabr'), (12, 'Dekabr')
        ],
        'years': range(today.year - 2, today.year + 3),
        'today': today,
        'stats': {
            'present': total_present,
            'absent': total_absent,
            'percent': overall_percentage,
        }
    })


@teacher_required
def session_create(request, group_pk):
    group = get_object_or_404(Group, pk=group_pk)
    if not request.user.is_admin_role and group.teacher.user != request.user:
        messages.error(request, "Faqat o'zingizning guruhlaringiz uchun davomat olishingiz mumkin!")
        return redirect('attendance:session_list')

    students = Enrollment.objects.filter(group=group, is_active=True).select_related('student')

    if request.method == 'POST':
        date_str = request.POST.get('date', str(timezone.now().date()))
        topic = request.POST.get('topic', '')
        session, _ = AttendanceSession.objects.get_or_create(
            group=group, date=date_str,
            defaults={'topic': topic, 'created_by': request.user}
        )

        for enrollment in students:
            status = request.POST.get(f'status_{enrollment.student.pk}', 'absent')
            notes = request.POST.get(f'notes_{enrollment.student.pk}', '')
            Attendance.objects.update_or_create(
                session=session,
                student=enrollment.student,
                defaults={'status': status, 'notes': notes}
            )
            
            if status == 'absent' and enrollment.student.telegram_id and enrollment.student.telegram_notifications:
                msg = (
                    f"❗ <b>Darsda qatnashmadi</b>\n\n"
                    f"O'quvchi: {enrollment.student.get_full_name()}\n"
                    f"Guruh: {group.name}\n"
                    f"Sana: {date_str}\n\n"
                    f"Iltimos, darslarni qoldirmaslikka harakat qiling."
                )
                send_telegram_message(enrollment.student.telegram_id, msg)

            if status == 'absent' and enrollment.student.user:
                Notification.objects.create(
                    recipient=enrollment.student.user,
                    sender=request.user,
                    title="Darsda qatnashmadi",
                    message=f"Siz {date_str} kuni {group.name} guruhidagi darsda qatnashmadi deb belgilandingiz."
                )

        messages.success(request, f"{date_str} sanasi uchun davomat saqlandi!")
        return redirect('attendance:session_list')

    today = str(timezone.now().date())
    return render(request, 'attendance/session_create.html', {
        'group': group,
        'students': students,
        'today': today,
        'statuses': [s for s in Attendance.Status.choices if s[0] in ['present', 'absent']],
    })


@login_required
def session_detail(request, pk):
    session = get_object_or_404(
        AttendanceSession.objects.select_related('group__course', 'created_by'), pk=pk
    )
    records = session.records.select_related('student').order_by('student__last_name')
    present_count = records.filter(status=Attendance.Status.PRESENT).count()
    total_count = records.count()
    return render(request, 'attendance/session_detail.html', {
        'session': session,
        'records': records,
        'present_count': present_count,
        'total_count': total_count,
    })


@teacher_required
def session_delete(request, pk):
    session = get_object_or_404(AttendanceSession, pk=pk)
    if not request.user.is_admin_role and session.group.teacher.user != request.user:
        messages.error(request, "Sizda ushbu sessiyani o'chirish huquqi yo'q!")
        return redirect('attendance:session_list')
    if request.method == 'POST':
        session.delete()
        messages.success(request, 'Sessiya o\'chirildi!')
    return redirect('attendance:session_list')


@require_POST
@login_required
@teacher_required
def attendance_update_ajax(request):
    try:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        date_str = data.get('date')
        status = data.get('status')
        group_id = data.get('group_id')
        
        try:
            date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return JsonResponse({'status': 'error', 'message': 'Noto\'g\'ri sana'}, status=400)
        
        if date_obj > timezone.now().date():
            return JsonResponse({'status': 'error', 'message': 'Kelajakdagi sanaga davomat qilib bo\'lmaydi'}, status=400)

        session, _ = AttendanceSession.objects.get_or_create(
            group_id=group_id, date=date_obj,
            defaults={'created_by': request.user}
        )
        
        if not request.user.is_admin_role and session.group.teacher.user != request.user:
            return JsonResponse({'status': 'error', 'message': 'Ruxsat yo\'q'}, status=403)
        
        if status in ['present', 'absent']:
            att, _ = Attendance.objects.update_or_create(
                session=session, student_id=student_id,
                defaults={'status': status}
            )
            
            # Non-blocking notification tasks
            try:
                if status == 'absent' and att.student.telegram_id and att.student.telegram_notifications:
                    msg = (
                        f"❗ <b>Darsda qatnashmadi</b>\n\n"
                        f"O'quvchi: {att.student.get_full_name()}\n"
                        f"Guruh: {session.group.name}\n"
                        f"Sana: {session.date}\n\n"
                        f"Iltimos, darslarni qoldirmaslikka harakat qiling."
                    )
                    send_telegram_message(att.student.telegram_id, msg)
                
                if status == 'absent' and att.student.user:
                    Notification.objects.get_or_create(
                        recipient=att.student.user,
                        sender=request.user,
                        title="Darsda qatnashmadi",
                        message=f"Siz {session.date} kuni {session.group.name} guruhidagi darsda qatnashmadi deb belgilandingiz."
                    )
            except Exception as e:
                print(f"Notification error: {e}")
        else:
            Attendance.objects.filter(session=session, student_id=student_id).delete()
            
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_POST
@login_required
@teacher_required
def update_topic_ajax(request):
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        date_str = data.get('date')
        topic = data.get('topic', '').strip()

        try:
            date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return JsonResponse({'status': 'error', 'message': 'Noto\'g\'ri sana formati'}, status=400)

        session, _ = AttendanceSession.objects.get_or_create(
            group_id=group_id, date=date_obj,
            defaults={'created_by': request.user}
        )

        if not request.user.is_admin_role and session.group.teacher.user != request.user:
            return JsonResponse({'status': 'error', 'message': 'Ruxsat yo\'q'}, status=403)

        session.topic = topic
        session.save(update_fields=['topic'])
        return JsonResponse({'status': 'success', 'topic': session.topic})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
@teacher_required
def export_attendance_excel(request):
    group_id = request.GET.get('group')
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    today = timezone.now().date()
    try:
        month = int(month) if month else today.month
        year = int(year) if year else today.year
    except ValueError:
        month, year = today.month, today.year
        
    if not group_id:
        return redirect('attendance:session_list')
        
    group = get_object_or_404(Group, pk=group_id)
    if not request.user.is_admin_role and group.teacher.user != request.user:
        messages.error(request, "Ruxsat yo'q!")
        return redirect('attendance:session_list')

    start_date = datetime.date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = datetime.date(year, month, last_day)
    
    potential_dates = [start_date + datetime.timedelta(days=i) for i in range(last_day)]
    schedule = group.days
    if schedule == Group.DayChoices.ODD:
        all_dates = [d for d in potential_dates if d.weekday() in [0, 2, 4]]
    elif schedule == Group.DayChoices.EVEN:
        all_dates = [d for d in potential_dates if d.weekday() in [1, 3, 5]]
    elif schedule == Group.DayChoices.WEEKEND:
        all_dates = [d for d in potential_dates if d.weekday() in [5, 6]]
    else:
        all_dates = potential_dates

    sessions_qs = AttendanceSession.objects.filter(
        group=group, date__range=[start_date, end_date]
    ).values('date', 'topic')
    session_topic_map = {s['date']: s['topic'] or '' for s in sessions_qs}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Davomat"
    
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_dates) + 2)
    ws['A1'] = f"{group.name} Davomati ({start_date.strftime('%B %Y')})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.cell(row=2, column=1, value="Dars Mavzusi").font = Font(bold=True)
    for col_idx, d in enumerate(all_dates, 2):
        topic = session_topic_map.get(d, "")
        ws.cell(row=2, column=col_idx, value=topic).alignment = Alignment(wrap_text=True, vertical='center')
    
    headers = ['F.I.SH'] + [d.strftime('%d') for d in all_dates] + ['Jami (B/Y/%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    enrollments = Enrollment.objects.filter(group=group, is_active=True).select_related('student').order_by('student__last_name')
    attendance_records = Attendance.objects.filter(
        session__group=group, session__date__range=[start_date, end_date]
    ).values('student_id', 'session__date', 'status')
    
    attendance_map = {}
    for r in attendance_records:
        sid, d, s = r['student_id'], r['session__date'], r['status']
        if sid not in attendance_map: attendance_map[sid] = {}
        attendance_map[sid][d] = s
        
    row_idx = 4
    for e in enrollments:
        ws.cell(row=row_idx, column=1, value=str(e.student))
        present_count = 0
        absent_count = 0
        for col_offset, d in enumerate(all_dates, 2):
            status = attendance_map.get(e.student.pk, {}).get(d)
            val = ''
            if status == 'present':
                val = 'Bor'
                present_count += 1
            elif status == 'absent':
                val = "Yo'q"
                absent_count += 1
            ws.cell(row=row_idx, column=col_offset, value=val).alignment = Alignment(horizontal='center')
        
        total_sessions = present_count + absent_count
        percent = (present_count / total_sessions * 100) if total_sessions > 0 else 0
        summary_text = f"{present_count} / {absent_count} / {int(percent)}%"
        ws.cell(row=row_idx, column=len(all_dates) + 2, value=summary_text).alignment = Alignment(horizontal='center')
        row_idx += 1
        
    for col_idx in range(1, len(all_dates) + 3):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="davomat_{group.id}_{start_date.month}_{start_date.year}.xlsx"'
    wb.save(response)
    return response
